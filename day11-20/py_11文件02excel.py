"""
author :rain
Date : 2021/01/20
Description :
"""
import csv
from time import sleep

import xlwt
import xlrd
import openpyxl
from openpyxl import Workbook
import datetime
import PyPDF2


def test01():
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Data can be assigned directly to cells
    ws['A1'] = 42
    # Rows can also be appended
    ws.append([1, 2, 3])
    # Python types will automatically be converted
    ws['A2'] = datetime.datetime.now()
    # Save the file
    wb.save("../filedir/sample.xlsx")


def test02():
    # sheets = wb.sheet_names()
    wb = openpyxl.load_workbook("../filedir/饭角测试用例全2.0.xlsx")
    # wb = openpyxl.open("../filedir/饭角测试用例全2.0.xlsx")
    # 获取sheet名
    print(wb.sheetnames)
    # 第一个sheet页的 行数 和 列数
    sheet1 = wb.worksheets[0]
    print(sheet1.max_row, sheet1.max_column)
    # 创建sheet页里的内容遍历对象并遍历
    generator_rows = sheet1.iter_rows(0, 3, 0, 6, True)
    generator_cols = sheet1.iter_cols(0, 6, 0, 3, True)
    for i in generator_rows:
        print(i)
    print('_' * 100)
    for i in generator_cols:
        print(i)
    # 修改表格中的内容
    sheet1.cell(2, 1).value = 'hahahahahahahahhaha'
    # cell1.value = 'testvalue'
    # cell1.comment = 'testcomment'
    wb.save("../filedir/饭角测试用例全2.0.xlsx")
    wb.close()


def test03():
    # 整理复制报销表格
    bx = []
    with open('../filedir/baoxiao1.csv', 'r+') as file:
        reader = csv.reader(file)
        for line in reader:
            i = reader.line_num
            # print(i, end='\t')
            for elem in line:
                if i % 7 in [2, 4, 6, 0]:
                    bx.append(elem)
        print('_' * 100)
        print(len(bx))
    for a in bx:
        print(a)
    wb = openpyxl.load_workbook("../filedir/高雨测试退款订单明细_饭角226web充值测试.xlsx")
    sheet1 = wb.worksheets[0]
    # for i in range(int(len(bx) / 4) + 1):
    l = 3
    for i in range(len(bx)):
        if (i + 1) % 4 == 1:
            sheet1.cell(l, 6).value = bx[i]
        if (i + 1) % 4 == 2:
            sheet1.cell(l, 3).value = bx[i]
        if (i + 1) % 4 == 3:
            sheet1.cell(l, 4).value = bx[i]
        if (i + 1) % 4 == 0:
            sheet1.cell(l, 5).value = bx[i]
            l += 1
    wb.save("../filedir/高雨测试退款订单明细_饭角226web充值测试.xlsx")
    wb.close()


if __name__ == '__main__':
    # test02()
    test03()
